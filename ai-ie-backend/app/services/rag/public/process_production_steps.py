import codecs
import datetime
from loguru import logger
import re
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Tuple, Optional


def parse_process_json(raw_text: str,xbkh: str) -> Tuple[str, List[Dict[str, Any]]]:
    """
    解析工序数据（包含款号对象 + 工序列表 + 待确认事项）

    Args:
        raw_text: 包含 JSON 代码块和额外文本的原始字符串

    Returns:
        Tuple[str, List[Dict[str, Any]]]: (款号, 工序列表)
    """
    try:
        # 1. 提取 JSON 代码块
        # 使用正则查找 ```json 和 ``` 之间的内容
        json_match = re.search(r'```(?:json)?\s*([\s\S]*?)\s*```', raw_text)
        # if json_match:
        #     # 【关键修改】使用 .group(1) 获取捕获组的内容
        #     json_str = json_match.group(1)
        #
        #     # 打印完整的 JSON 字符串
        #     logger.info(json_str)
        # else:
        #     logger.error("未找到 JSON 数据")
        # if not json_match:
        #     # 如果没有 markdown 标记，尝试直接解析整个字符串（容错）
        #     # 或者查找第一个 [ 和最后一个 ]
        #     start = raw_text.find('[')
        #     end = raw_text.rfind(']')
        #     if start != -1 and end != -1:
        #         json_str = raw_text[start:end + 1]
        #     else:
        #         raise ValueError("未找到有效的 JSON 数据块")
        # else:
        json_str = json_match.group(1).strip()
        logger.info(f"提取后的JSON开头: {json_str[:50]}...")
        # json_str_d = codecs.decode(json_str, 'unicode_escape')
        # 2. 解析 JSON 为 Python 列表
        data_list = json.loads(json_str)
        logger.info(data_list)
        if not isinstance(data_list, list):
            raise ValueError("JSON 根节点必须是列表")

        # 3. 提取款号 (通常在第一个对象中)
        style_no = xbkh


        # for item in data_list:
        #     # 检查是否包含款号字段
        #     if "款号" in item:
        #         style_no = item["款号"]


        # 4. 数据清洗与标准化 (可选)
        # 确保序号是整数，工价是浮点数等
        clean_steps = []
        for step in data_list[1:]:
            clean_step = {
                "序号": int(step.get("序号", 0)),
                "流程阶段": step.get("流程阶段", ""),
                "工序名称": step.get("工序名称", ""),
                "操作说明": step.get("操作说明", ""),
                "设备工具": step.get("设备工具", "")
            }
            clean_steps.append(clean_step)


        # 按序号排序（防止 AI 生成时顺序错乱）
        clean_steps.sort(key=lambda x: x["序号"])
        logger.info(f"成功解析新版工序JSON | 款号={style_no} | 工序数量={len(clean_steps)}")
        return style_no, clean_steps

    except json.JSONDecodeError as e:
        logger.error(f"JSON解析失败: {str(e)} | 原始数据片段={json_str[:100]}...")
        raise ValueError(f"无效的JSON格式: {str(e)}")
    except Exception as e:
        logger.exception(f"解析工序数据时出错: {str(e)}")
        raise


def generate_excel(style_no: str, steps: List[Dict[str, Any]], person_name: str) -> bytes:
    """
    生成包含工序信息的Excel文件
    包含字段：序号、流程阶段、工序名称、操作说明、设备工具
    """
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "工序表"

    # --- 样式定义 ---
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # 第一行：标题 (合并 A1:E1)
    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value="箱包工序流程表")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center")

    # 第二行：款号 (合并 B2:E2)
    ws.cell(row=2, column=1, value="款号")
    ws.cell(row=2, column=2, value=style_no)
    ws.merge_cells('B2:E2')

    # 第三行：时间和人员
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.cell(row=3, column=1, value="时间")
    ws.cell(row=3, column=2, value=current_time)
    ws.cell(row=3, column=4, value="人员")
    ws.cell(row=3, column=5, value=person_name)
    # 合并 B3:C3 给时间，D3:E3 给人员（或者保持原逻辑，这里稍微调整以适应5列布局）
    ws.merge_cells('B3:C3')

    # 第四行：表头 (严格按照你提供的字段)
    headers = ["序号", "流程阶段", "工序名称", "操作说明", "设备工具"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 写入工序数据
    for row_idx, step in enumerate(steps, 5):  # 从第5行开始
        ws.cell(row=row_idx, column=1, value=int(step.get("序号", 0)))
        ws.cell(row=row_idx, column=2, value=step.get("流程阶段", ""))
        ws.cell(row=row_idx, column=3, value=step.get("工序名称", ""))
        ws.cell(row=row_idx, column=4, value=step.get("操作说明", ""))
        ws.cell(row=row_idx, column=5, value=step.get("设备工具", ""))

    # 设置列宽 (根据内容调整)
    ws.column_dimensions['A'].width = 8  # 序号
    ws.column_dimensions['B'].width = 15  # 流程阶段
    ws.column_dimensions['C'].width = 20  # 工序名称
    ws.column_dimensions['D'].width = 40  # 操作说明 (通常内容较长)
    ws.column_dimensions['E'].width = 20  # 设备工具

    # 设置边框
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # 应用边框到所有数据单元格
    max_row = len(steps) + 4
    max_col = 5

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            # 可选：给操作说明列设置自动换行
            if col == 4:
                cell.alignment = Alignment(wrap_text=True)

    # 将工作簿保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
